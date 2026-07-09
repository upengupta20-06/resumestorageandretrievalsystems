"""
Export helpers: build an in-memory Excel workbook or PDF table of candidates.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

HEADERS = ["Name", "Email", "Phone", "Department", "Skills", "Experience (yrs)", "Status", "Upload Date"]


def _row_for(resume):
    return [
        resume.get("name", ""),
        resume.get("email", ""),
        resume.get("phone", ""),
        resume.get("department", ""),
        ", ".join(resume.get("skills", [])),
        resume.get("experience_years", ""),
        resume.get("status", ""),
        resume.get("upload_date").strftime("%Y-%m-%d") if resume.get("upload_date") else "",
    ]


def export_to_excel(resumes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"

    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for resume in resumes:
        ws.append(_row_for(resume))

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_to_pdf(resumes):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = [Paragraph("Candidate List Export", styles["Title"]), Spacer(1, 12)]

    data = [HEADERS] + [_row_for(r) for r in resumes]
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3B82F6")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer
